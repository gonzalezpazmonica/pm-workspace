#!/usr/bin/env python3
"""
generate_fixtures.py — Deterministic XLSX fixtures for SE-324.

Creates the .xlsx fixtures used by tests/bats/test-se324-tabular-excel.bats.
Requires openpyxl (optional dependency of tabular-profile.py). The output is
committed so the tests can run even in environments where openpyxl is not
installed at generation time.

Fixtures:
- ventas_one.xlsx   -> 1 sheet "ventas" (mirror of ventas.csv)  [AC-1.3]
- ventas_multi.xlsx -> 3 sheets: ventas, clientes, productos   [AC-1.1, AC-2.x]
- formulas.xlsx     -> formula cells WITH cached computed values [AC-1.2]

Formulas with cached values are produced by writing the workbook with openpyxl
and then injecting the cached <v> elements into the sheet XML — mimicking how
Excel saves a workbook (formula + last-computed result). openpyxl itself does
not compute formulas, so without this step data_only=True would return None.
"""
import re
import sys
import zipfile
from pathlib import Path

import openpyxl  # noqa: E402

HERE = Path(__file__).resolve().parent

VENTAS = [
    ["id", "producto", "cantidad", "precio", "importe"],
    [1, "Teclado", 10, 25.0, 250.0],
    [2, "Raton", 15, 12.5, 187.5],
    [3, "Monitor", 5, 150.0, 750.0],
    [4, "Webcam", 8, 45.0, 360.0],
    [5, "Altavoz", 12, 30.0, 360.0],
    [6, "Cable", 20, 3.5, 70.0],
]

CLIENTES = [
    ["id_cliente", "nombre"],
    [1, "Ana"],
    [2, "Luis"],
    [3, "Carla"],
    [4, "Diego"],
    [5, "Eva"],
]

PRODUCTOS = [
    ["id_producto", "nombre", "precio"],
    [1, "Teclado", 25.0],
    [2, "Raton", 12.5],
    [3, "Monitor", 150.0],
    [4, "Webcam", 45.0],
    [5, "Altavoz", 30.0],
    [6, "Cable", 3.5],
]


def write_sheet(ws, data):
    for row in data:
        ws.append(row)


def make_ventas_one(path):
    wb = openpyxl.Workbook()
    write_sheet(wb.active, VENTAS)
    wb.active.title = "ventas"
    wb.save(path)


def make_ventas_multi(path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "ventas"
    write_sheet(ws1, VENTAS)
    ws2 = wb.create_sheet("clientes")
    write_sheet(ws2, CLIENTES)
    ws3 = wb.create_sheet("productos")
    write_sheet(ws3, PRODUCTOS)
    wb.save(path)


def make_formulas(path):
    """Write formulas.xlsx and inject cached computed values into sheet XML."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "formulas"
    ws.append(["operacion", "a", "b", "resultado"])
    ws["A2"] = "suma"
    ws["B2"] = 10
    ws["C2"] = 5
    ws["D2"] = "=B2+C2"        # -> 15
    ws["A3"] = "producto"
    ws["B3"] = 7
    ws["C3"] = 3
    ws["D3"] = "=B3*C3"        # -> 21
    ws["A4"] = "resta"
    ws["B4"] = 100
    ws["C4"] = 4
    ws["D4"] = "=B4-C4"        # -> 96
    wb.save(path)

    cached = {
        "D2": "15",
        "D3": "21",
        "D4": "96",
    }
    tmp = path.with_suffix(".tmp")
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w",
                                                       zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                xml = data.decode("utf-8")
                for cell, value in cached.items():
                    # openpyxl writes formula cells as
                    #   <c r="D2"><f>B2+C2</f><v /></c>
                    # Replace the empty <v /> with the cached computed value,
                    # mimicking how Excel persists formula results (AC-1.2).
                    pattern = re.compile(
                        r'(<c r="%s"[^>]*><f>[^<]*</f>)<v\s*/>' % cell
                    )
                    xml = pattern.sub(
                        r"\1<v>%s</v>" % value, xml
                    )
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    tmp.replace(path)


def main():
    make_ventas_one(HERE / "ventas_one.xlsx")
    make_ventas_multi(HERE / "ventas_multi.xlsx")
    make_formulas(HERE / "formulas.xlsx")
    print("fixtures generated:", HERE)
    return 0


if __name__ == "__main__":
    sys.exit(main())
